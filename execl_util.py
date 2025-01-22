import os
import shutil
from openpyxl import load_workbook, Workbook


# 将带有 "万" 的字符串转换为数字，仅对 BgCg_ebQ 列处理
def convert_to_number(value):
    if isinstance(value, str) and '万' in value:
        value = value.replace('万', '')  # 去掉 "万"
        return float(value) * 10000  # 转换为实际的数字，乘以 10000
    try:
        return float(value)  # 如果是数值，直接转换为浮动数
    except ValueError:
        return value  # 如果无法转换为数字，返回原始值


def sort_excel(file_path):
    # 读取 Excel 文件
    wb = load_workbook(file_path)
    ws = wb.active

    # 获取行数和列数
    total_rows = ws.max_row
    total_cols = ws.max_column

    # 获取表头
    headers = [ws.cell(1, col).value for col in range(1, total_cols + 1)]

    # 找到 BgCg_ebQ 列的索引
    bgcg_ebq_col_index = -1
    for col_index, header in enumerate(headers):
        if header == "BgCg_ebQ":
            bgcg_ebq_col_index = col_index
            break

    if bgcg_ebq_col_index == -1:
        print("没有找到 'BgCg_ebQ' 列。")
        return

    # 获取 BgCg_ebQ 列的数据并转换
    data = []
    for row in range(2, total_rows + 1):
        row_data = [ws.cell(row, col).value for col in range(1, total_cols + 1)]
        # 将 BgCg_ebQ 列转换为数值
        row_data[bgcg_ebq_col_index] = convert_to_number(row_data[bgcg_ebq_col_index])
        data.append(row_data)

    # 按照 BgCg_ebQ 列进行排序，保证整行数据不变
    data.sort(key=lambda x: x[bgcg_ebq_col_index], reverse=True)

    # 创建一个新的工作簿，并添加表头
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.append(headers)

    # 写入排序后的数据
    for row_data in data:
        new_ws.append(row_data)

    # 保存排序后的文件
    sorted_file_path = 'sorted_' + os.path.basename(file_path)
    new_wb.save(sorted_file_path)
    print(f"Sorted Excel file saved as {sorted_file_path}")

    return sorted_file_path  # 返回排序后的文件路径


def split_excel(file_path, output_dir, chunk_size):
    # 读取排序后的 Excel 文件
    wb = load_workbook(file_path)
    ws = wb.active

    # 获取行数和列数
    total_rows = ws.max_row
    total_cols = ws.max_column

    # 获取表头
    headers = [ws.cell(1, col).value for col in range(1, total_cols + 1)]

    # 如果输出目录存在，先清空目录
    if os.path.exists(output_dir):
        shutil.rmtree(output_dir)
    # 重新创建输出目录
    os.makedirs(output_dir)

    # 按 chunk_size 拆分数据
    for i in range(0, total_rows - 1, chunk_size):
        new_wb = Workbook()
        new_ws = new_wb.active

        # 写入表头
        new_ws.append(headers)

        # 写入每块数据
        for row in range(i + 2, min(i + chunk_size + 2, total_rows + 1)):
            new_row = [ws.cell(row, col).value for col in range(1, total_cols + 1)]
            new_ws.append(new_row)

        # 保存文件到指定目录
        output_file = os.path.join(output_dir, f'output_part_{i // chunk_size + 1}.xlsx')
        new_wb.save(output_file)
        print(f'Saved: {output_file}')


def main():
    # 输入文件路径和拆分行数
    file_path = 'douyin29.xlsx'  # 替换为你的Excel文件路径
    output_dir = 'execl_directory'  # 替换为你希望保存文件的输出目录
    chunk_size = 98  # 每个新文件的行数，可以根据需求更改
 
    # 1. 排序文件并保存排序后的文件
    sorted_file_path = sort_excel(file_path)

    # 2. 调用拆分函数
    split_excel(sorted_file_path, output_dir, chunk_size)


if __name__ == '__main__':
    main()
